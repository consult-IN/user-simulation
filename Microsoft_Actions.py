import os
import configparser

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import random
import english_words

import win32com.client
from datetime import datetime, timedelta


   
def downloading_Attachements():
# Downloadet Anhänge von Mails in Outlook
    config = configparser.ConfigParser()
    config.read('ConfigPathDriver.txt')

   
    outputdir = r"C:\Users\\"+config['PATH']['User'] +r"\Email Anhänge"

# startet die Sitzung und holt sich per MAPI den zugriff auf die emails
    outlook = win32com.client.Dispatch('outlook.application')
    mapi = outlook.GetNamespace("MAPI")

    # erhält so die Emails im Ordner Default

    inbox = mapi.GetDefaultFolder(6)
    
# holt sich den Text und Anhänge der Emails
    messages = inbox.Items
    message = messages.GetFirst()
    i = 0
    try:
        for mess in list(messages) :
            try:
                message = messages.GetNext()
                attachments = message.Attachments
                attach = attachments.Item(1)
                
                for attach in message.Attachments:
                    attach.SaveAsFile(os.path.join(outputdir,str(attach)))

            except:
                pass
    except:
        pass


def createExcel():
# Erstellt Excel-Datei, benennt Tabelle, fügt einer Reihe Werte hinzu

   valueCell1 = random.random()
   valueCell2 = random.random()
   valueCell3 = random.random()
   valueCell4 = random.random()

   wordRandom = english_words.english_words_set.pop()

   wb = Workbook()  # creates workbook
   ws = wb.active  # creates default worksheet
   ws.title = wordRandom  # Name worksheet

   ws.append([valueCell1, valueCell2, valueCell3, valueCell4])  # apppend new values in a row

   ws.merge_cells("C1:D1")

   wb.save('Great.xlsx')

def existingExcelFile():
# Kreiert neue Tabelle in vorhandener Excel-Datei, verändert Werte einer Zelle

   randomWord = english_words.english_words_set.pop()
   wb = load_workbook('Great.xlsx')

   wb.create_sheet(randomWord)  # create new sheet
   ws = wb.active # shows active worksheet


   #print(ws['A1'].value)  # print the value in a cell

   ws['A2'].value = randomWord  # change value in a cell

   # save changes
   wb.save('Great.xlsx')
   
def accessMultipleExcelCells():
# Benennt und verbindet Zellen, fügt Zeilen und Spalten in eingegebener Reihe ein, löscht diese und verschiebt Zellen
   
   valueNum = random.randint(1, 5)
   valueNum2 = random.randint(6, 15)
   wb = load_workbook('Great.xlsx')
   ws = wb.active

   for row in range(valueNum, valueNum2):  # look at rows x-y
       for col in range(valueNum, valueNum2):  # look at columns x-y
           char = get_column_letter(col)  # gives column letters per number
           ws[char + str(row)] = char + str(row)  # renames cells (In this example: Corresponding cell name)
           ws[char + str(row)].value
          
   
   ws.insert_rows(valueNum2)  # insert row in row int
   ws.delete_rows(valueNum)  # delete row in row int

   ws.insert_cols(valueNum2)
   ws.delete_cols(valueNum)

   ws.move_range("C1:D11", rows=2, cols=2)  # move cells ranging from C1 to D11, rows(negative int -> up) cols (negative int -> left)

   wb.save('Great.xlsx')


def wordToPdf():
# Word Dateien werden in pdf Dateien umgewandelt

    # Enter existing word document name
    wordFilename = "Enter_your_WordFileName.docx"
    # Enter how ypu want to name your pdf file
    pdfFilename = "Enter_your_PDFFileName.pdf"

    doc = docx.Document()
    # Code to create Word document goes here
    doc.save(wordFilename)

    wdFormatPDF = 17  # Word's numeric code for PDFs.
    wordObj = win32com.client.Dispatch('Word.Application')

    docObj = wordObj.Documents.Open(wordFilename)
    docObj.SaveAs(pdfFilename, FileFormat=wdFormatPDF)
    docObj.Close()
    wordObj.Quit()
   

def randomizer_Microsoft():
    a = randint(1,5)
    if a == 1:
        downloading_Attachements()
        
    if a == 2:
        createExcel()
        
    if a == 3:
        existingExcelFile()
       
    if a == 4:
        accessMultipleExcelCells()
        
    if a == 5:
        wordToPdf()
        
